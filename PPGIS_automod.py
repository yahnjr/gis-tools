from arcgis import features
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

# Check if the Excel file exists and delete it to allow overwriting
excel_file_path = ****output_file_path
if os.path.exists(excel_file_path):
    os.remove(excel_file_path)

# Initialize a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Replace with the URL of your public AGOL shapefile layer
url = ***url_comments_layer

# Replace with the name of the field containing comments
comments_field = "pubcomment"

# List of bad words (currently only contains "the" for testing)
bad_words = [***list_of_bad_words]

# Counter for comments with bad words
bad_comment_count = 0

# Access the shapefile layer
layer = features.FeatureLayer(url)

# Excel row start
row = 1

# Loop through each feature in the layer
query_result = layer.query()  # Store query result to ensure all features are processed
for feature in query_result:
    # Ensure comments attribute exists and is not None before proceeding
    if (
        comments_field in feature.attributes
        and feature.attributes[comments_field] is not None
    ):
        # Get the comments attribute value
        comments = feature.attributes[
            comments_field
        ].lower()  # Convert to lowercase for case-insensitive search

        # Write comments to Excel, check for bad words and apply styles if found
        ws[f"A{row}"] = comments
        for bad_word in bad_words:
            if bad_word in comments:
                bad_comment_count += 1
                # Apply red fill and bold font to cells with bad words
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"A{row}"].fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                break  # Exit inner loop after finding a bad word
        row += 1

# Save the workbook
wb.save(excel_file_path)

# Create the alert message
message = f"There are {bad_comment_count} comments that need to be reviewed."

# Simulate an alert by printing the message to the console (you can replace this with your preferred alert method)
print(message)

import os

# ... your existing code ...

print(bad_comment_count)

# Display a pop-up message (Windows only)
if bad_comment_count > 0:
    os.system(f"msg * {message}")
